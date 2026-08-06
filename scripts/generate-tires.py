#!/usr/bin/env python
"""
Generate artifacts/aeolus-website/src/data/tires.generated.ts from the
"3. Content Wireframe" sheet of Aeolus-Wireframe-06.xlsx.

Sheet 3 is the content bible: every tire in it must appear on the site, and
nothing on the site should contradict it. Sheets 1 and 2 are internal notes
and are deliberately NOT read here.

Usage:  pnpm --filter @workspace/scripts run generate:tires
Requires: openpyxl  (pip install openpyxl)
"""

import os
import re
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  python -m pip install openpyxl")

try:
    from PIL import Image
except ImportError:
    sys.exit("Pillow is required:  python -m pip install Pillow")

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WORKBOOK = os.path.join(ROOT, "Aeolus-Wireframe-06.xlsx")
SHEET = "3. Content Wireframe"
SITE = os.path.join(ROOT, "artifacts", "aeolus-website")
OUT = os.path.join(SITE, "src", "data", "tires.generated.ts")
PHOTO_DIR = os.path.join(SITE, "public", "tires", "Tire-Photos")
FEATURE_DIR = os.path.join(SITE, "public", "tires", "Feature-Images")

# ── Column map for the repeating spec block ──────────────────────────────────
COLS = {
    "size": 5, "ply": 6, "rimW": 7, "secW": 8, "odIn": 9, "odMm": 10,
    "tdMm": 11, "td32": 12,
    "mlSlbs": 13, "mlSpsi": 14, "mlSkg": 15, "mlSkpa": 16,
    "mlDlbs": 17, "mlDpsi": 18, "mlDkg": 19, "mlDkpa": 20,
    "liss": 21,
}
FLAG_COLS = {"smartway": 22, "ms": 23, "3PMSF": 24}

# ── Wireframe category → Navbar dropdown section ─────────────────────────────
NAV_GROUP = {
    "Premium Long Haul":   "NEO SERIES LONG HAUL",
    "Premium Regional":    "NEO/SAILOR SERIES REGIONAL",
    "Premium On/Off Road": "NEO SERIES ON/OFF ROAD",
    "Premium Winter":      "NEO SERIES WINTER",
    "Premium Urban":       "NEO SERIES URBAN",
    "Standard Long Haul":  "STANDARD SERIES LONG HAUL",
    "Standard Regional":   "STANDARD SERIES REGIONAL",
    "Standard On/Off":     "STANDARD SERIES ON/OFF ROAD",
    "Standard Off Road":   "STANDARD SERIES ON/OFF ROAD",
    "Standard Winter":     "STANDARD SERIES WINTER",
}

# Canonical TirePosition values (must match tire-types.ts). The workbook's
# casing sometimes drifts ("All position" vs "All Position") — match
# case-insensitively so the site's POS_SVG icon lookup doesn't silently fail.
POSITIONS = ["Drive", "Steer", "Trailer", "All Position", "OTR", "Bus", "Steer/Trailer"]
POSITION_LOOKUP = {p.lower(): p for p in POSITIONS}

# Order the dropdown columns render in.
NAV_ORDER = [
    "NEO SERIES LONG HAUL", "NEO/SAILOR SERIES REGIONAL", "NEO SERIES ON/OFF ROAD",
    "NEO SERIES WINTER", "NEO SERIES URBAN",
    "STANDARD SERIES LONG HAUL", "STANDARD SERIES REGIONAL",
    "STANDARD SERIES ON/OFF ROAD", "STANDARD SERIES WINTER",
]

# Photo filename stems that don't follow "name with spaces → hyphens".
PHOTO_OVERRIDES = {
    "Neo Construct G": "Aeolus-Neo-Construct-G",
}

WEBP_QUALITY = 90

# ── Typo repair ──────────────────────────────────────────────────────────────
# The workbook is missing a space before "and" in 47 places ("uniform wearand
# excellent handling"). Every affected token is listed explicitly rather than
# regex-matched, so real words like "brand"/"stand"/"demand" can never be hit.
AND_TYPOS = [
    "wearand", "tractionand", "resistanceand", "noiseand", "integrityand",
    "lifespanand", "controland", "rideand", "comfortand", "rejectionand",
    "compoundand", "durableand", "stabilityand", "longevityand", "expulsionand",
    "lifeand", "safetyand", "shouldersand", "durabilityand", "lossesand",
    "heatand",
]
AND_RE = re.compile(r"\b(" + "|".join(t[:-3] for t in AND_TYPOS) + r")and\b")

# Tags that are legitimately camel-cased, so the run-together check skips them.
CAMEL_OK = {"SmartWay", "AllRoads", "AllSeason"}

typo_log = []


def fix_text(s, tire, field):
    if not s:
        return s
    fixed, n = AND_RE.subn(r"\1 and", s)
    if n:
        typo_log.append((tire, field, n))
    return fixed


# ── Slug / photo helpers ─────────────────────────────────────────────────────
def slugify(name):
    s = name.strip().lower()
    s = s.replace("+", "-plus")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return re.sub(r"-+", "-", s).strip("-")


def resolve_photo(stem, folder):
    """Find `stem`'s image in `folder`, converting a raw PNG drop to WebP the
    first time it's seen so Mark's drop-a-PNG workflow needs no changes but
    served files stay small. Idempotent: re-running finds the .webp and skips
    conversion. Returns the /public-relative URL, or None if nothing matches.
    """
    webp_path = os.path.join(folder, stem + ".webp")
    if os.path.exists(webp_path):
        return "/tires/" + os.path.basename(folder) + "/" + stem + ".webp"

    png_path = os.path.join(folder, stem + ".png")
    if os.path.exists(png_path):
        im = Image.open(png_path)
        im.save(webp_path, "WEBP", quality=WEBP_QUALITY, method=6)
        os.remove(png_path)
        return "/tires/" + os.path.basename(folder) + "/" + stem + ".webp"

    return None


def photo_for(name):
    stem = PHOTO_OVERRIDES.get(name, name.strip().replace(" ", "-"))
    url = resolve_photo(stem, PHOTO_DIR)
    return (url, True) if url else (None, False)


def feature_image(token):
    """Wireframe writes bare tokens like 'Neo-Fuel-D3-f1'; resolve to a real file."""
    if not token:
        return "", True
    token = token.strip()
    for ext in (".jpg", ".jpeg", ".png"):
        if os.path.exists(os.path.join(FEATURE_DIR, token + ext)):
            return "/tires/Feature-Images/" + token + ext, True
    return "", False


# ── Parse the sheet ──────────────────────────────────────────────────────────
def cell(ws, r, c):
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()


def num(v):
    """Trim Excel float noise: 10.1181102362205 → 10.1, 24.0 → 24."""
    if v == "":
        return ""
    try:
        f = float(v)
    except ValueError:
        return v
    if abs(f - round(f)) < 1e-9:
        return str(int(round(f)))
    return f"{round(f, 2):g}"


notes = []


def parse():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    tires, cur, mode = [], None, None

    for r in range(1, ws.max_row + 1):
        label = cell(ws, r, 1).lower()
        name = cell(ws, r, 2)

        if label == "tire name" and name:
            cur = {
                "name": name, "row": r, "category": "", "subtitle": "", "tags": "",
                "pos": "", "alt": "", "bullets": [], "features": [], "sizes": [],
            }
            tires.append(cur)
            mode = None
            continue
        if cur is None:
            continue

        # Decide the row type from the size cell, NOT from column A. Column A is
        # a free-text margin used for review notes ("ai rec' 11.7"), and keying
        # off it silently drops annotated spec rows.
        size = cell(ws, r, 5)
        if size and size.lower() != "size":
            row = {k: num(cell(ws, r, c)) for k, c in COLS.items()}
            row["size"] = size
            for k, c in FLAG_COLS.items():
                row[k] = cell(ws, r, c) != ""
            cur["sizes"].append(row)
            if label:
                notes.append(f"{cur['name']} {size}: margin note — {cell(ws, r, 1)}")
            continue

        if label:
            simple = {
                "category": ("category", 4), "subtitle": ("subtitle", 4),
                "pos": ("pos", 4), "tags": ("tags", 3), "alt": ("alt", 3),
            }
            if label in simple:
                key, col = simple[label]
                cur[key] = cell(ws, r, col) or cell(ws, r, 4 if col == 3 else 3)
                mode = None
                continue
            if label == "bullet points":
                mode = "bullets"
                t = cell(ws, r, 4)
                if t:
                    cur["bullets"].append(t)
                continue
            m = re.match(r"(title|body|image)\s*(\d+)", label)
            if m:
                kind, idx = m.group(1), int(m.group(2))
                while len(cur["features"]) < idx:
                    cur["features"].append({"title": "", "body": "", "image": ""})
                cur["features"][idx - 1][kind] = cell(ws, r, 3) or cell(ws, r, 4)
                mode = "feature"
                continue
            if label == "feature:":
                mode = "feature"
                continue
            mode = None
            continue

        if mode == "bullets":
            t = cell(ws, r, 4)
            if t:
                cur["bullets"].append(t)

    return tires


# ── Emit ─────────────────────────────────────────────────────────────────────
def ts(s):
    return '"' + str(s).replace("\\", "\\\\").replace('"', '\\"') + '"'


def main():
    tires = parse()
    warnings = []
    out = []

    for t in tires:
        name = t["name"]
        slug = slugify(name)
        cat = t["category"]
        nav = NAV_GROUP.get(cat, "")
        if cat and not nav:
            warnings.append(f"{name}: category {cat!r} has no navGroup mapping")
        if not cat:
            warnings.append(f"{name}: no category in the wireframe — kept out of the nav "
                            f"dropdown and shown without a series/category label on the grid")

        series, category_label = "", ""
        if cat.startswith("Premium"):
            series, category_label = "PREMIUM", cat[len("Premium "):].upper()
        elif cat.startswith("Standard"):
            series, category_label = "STANDARD", cat[len("Standard "):].upper()

        pos = POSITION_LOOKUP.get(t["pos"].strip().lower(), t["pos"])
        if t["pos"] and pos not in POSITIONS:
            warnings.append(f"{name}: position {t['pos']!r} doesn't match a known "
                            f"position — its icon will be missing on the site")

        photo, ok = photo_for(name)
        if not ok:
            warnings.append(f"{name}: no tire photo — falling back to the shared placeholder")
            photo = None

        alt_photo = None
        if t["alt"]:
            alt_stem = t["alt"].strip()
            alt_photo = resolve_photo(alt_stem, PHOTO_DIR)
            if alt_photo:
                # The wireframe pairs some alts with a photo from a different
                # model; surface that rather than silently shipping it.
                stem = name.replace(" ", "-").replace("Neo-", "").lower()
                if stem.split("-")[0] not in t["alt"].lower():
                    warnings.append(
                        f"{name}: alt photo is {t['alt']!r} — check this is the right model")
            else:
                warnings.append(f"{name}: alt photo {alt_stem + '.png'!r} not found")

        feats = []
        for i, f in enumerate(t["features"]):
            img, found = feature_image(f["image"])
            if f["image"] and not found:
                warnings.append(f"{name}: feature {i+1} image {f['image']!r} not found")
            if not (f["title"] or f["body"] or img):
                continue
            feats.append({
                "title": fix_text(f["title"], name, f"feature {i+1} title"),
                "body": fix_text(f["body"], name, f"feature {i+1} body"),
                "image": img,
            })

        tags = [x.strip() for x in t["tags"].split(",") if x.strip()]
        bullets = [fix_text(b, name, "bullet") for b in t["bullets"]]
        subtitle = fix_text(t["subtitle"], name, "subtitle")

        if not t["sizes"]:
            warnings.append(f"{name}: no size rows in the wireframe")

        # Plausibility check on the inch/mm columns. Catches unit slips, e.g. a
        # section width entered in mm into the inches column.
        for row in t["sizes"]:
            for field, lo, hi, unit in (
                ("rimW", 5, 20, "in"), ("secW", 6, 20, "in"),
                ("odIn", 25, 55, "in"), ("odMm", 650, 1400, "mm"),
                ("tdMm", 8, 40, "mm"), ("td32", 10, 45, "32nds"),
            ):
                v = row.get(field, "")
                if v == "":
                    continue
                try:
                    f = float(v)
                except ValueError:
                    continue
                if not (lo <= f <= hi):
                    warnings.append(
                        f"{name} {row['size']}: {field} = {v} is outside the plausible "
                        f"{lo}–{hi} {unit} range — check the workbook")

        out.append({
            "slug": slug, "name": name, "segment": cat, "position": pos,
            "navGroup": nav, "seriesLabel": series, "categoryLabel": category_label,
            "subtitle": subtitle, "tags": tags, "bullets": bullets,
            "features": feats, "specRows": t["sizes"],
            "tireImage": photo, "altImage": alt_photo,
        })

    # Near-duplicate tags become two separate filter facets for one concept.
    # They can only be fixed in the workbook, so surface them rather than
    # silently normalising and drifting from the bible.
    seen = {}
    for t in out:
        for tag in t["tags"]:
            # A lowercase letter butted against an uppercase one mid-word almost
            # always means two tags ran together ("Long LifevM+S"). Brand names
            # that are legitimately camel-cased are exempt.
            if re.search(r"[a-z][A-Z]", tag) and tag not in CAMEL_OK:
                warnings.append(f"{t['name']}: tag {tag!r} looks like two tags run "
                                f"together — fix in the workbook")
            key = re.sub(r"[^a-z]", "", tag.lower()).rstrip("s")
            seen.setdefault(key, {}).setdefault(tag, []).append(t["name"])
    for key, variants in seen.items():
        if len(variants) > 1:
            detail = "; ".join(f"{v!r} ({', '.join(names)})" for v, names in variants.items())
            warnings.append(f"tag spelled inconsistently in the workbook — {detail}")

    lines = [
        "// ─────────────────────────────────────────────────────────────────────────────",
        "// AUTO-GENERATED — DO NOT EDIT BY HAND.",
        "//",
        f'// Source: Aeolus-Wireframe-06.xlsx, sheet "{SHEET}" (the content bible).',
        f"// Regenerate: pnpm --filter @workspace/scripts run generate:tires",
        f"// Last generated: {date.today().isoformat()}",
        "//",
        "// Hand-maintained companions: tire-types.ts, demo-tires.ts.",
        "// ─────────────────────────────────────────────────────────────────────────────",
        "",
        'import { TireData, TirePosition, SHARED_ASSETS } from "./tire-types";',
        "",
        f"/** {len(out)} tires, in wireframe order. */",
        "export const BIBLE_TIRES: TireData[] = [",
    ]

    for t in out:
        lines.append("  {")
        lines.append(f'    slug:     {ts(t["slug"])},')
        lines.append(f'    name:     {ts(t["name"])},')
        lines.append(f'    segment:  {ts(t["segment"])},')
        lines.append(f'    position: {ts(t["position"])} as TirePosition,')
        lines.append(f'    navGroup: {ts(t["navGroup"])},')
        lines.append(f'    seriesLabel:   {ts(t["seriesLabel"])},')
        lines.append(f'    categoryLabel: {ts(t["categoryLabel"])},')
        lines.append(f'    subtitle: {ts(t["subtitle"])},')

        if t["tags"]:
            lines.append("    tags: [" + ", ".join(ts(x) for x in t["tags"]) + "],")
        else:
            lines.append("    tags: [],")

        if t["bullets"]:
            lines.append("    bullets: [")
            for b in t["bullets"]:
                lines.append(f"      {ts(b)},")
            lines.append("    ],")
        else:
            lines.append("    bullets: [],")

        if t["features"]:
            lines.append("    features: [")
            for f in t["features"]:
                lines.append("      {")
                lines.append(f'        title: {ts(f["title"])},')
                lines.append(f'        body:  {ts(f["body"])},')
                lines.append(f'        image: {ts(f["image"])},')
                lines.append("      },")
            lines.append("    ],")
        else:
            lines.append("    features: [],")

        if t["specRows"]:
            lines.append("    specRows: [")
            for r in t["specRows"]:
                parts = [f'size:{ts(r["size"])}']
                for k in ("ply", "rimW", "secW", "odIn", "odMm", "tdMm", "td32",
                          "mlSlbs", "mlSpsi", "mlSkg", "mlSkpa",
                          "mlDlbs", "mlDpsi", "mlDkg", "mlDkpa", "liss"):
                    parts.append(f"{k}:{ts(r[k])}")
                for k in ("smartway", "ms"):
                    parts.append(f"{k}:{'true' if r[k] else 'false'}")
                parts.append(f'"3PMSF":{"true" if r["3PMSF"] else "false"}')
                lines.append("      { " + ", ".join(parts) + " },")
            lines.append("    ],")
        else:
            lines.append("    specRows: [],")

        img = ts(t["tireImage"]) if t["tireImage"] else "SHARED_ASSETS.placeholderPhoto"
        lines.append(f"    tireImage:    {img},")
        if t["altImage"]:
            lines.append(f'    altImage:     {ts(t["altImage"])},')
        lines.append("    heroBg:       SHARED_ASSETS.heroBg,")
        lines.append("    bgTruck:      SHARED_ASSETS.bgTruck,")
        lines.append("    cutawayImage: SHARED_ASSETS.cutaway,")
        lines.append("    downloads: {")
        lines.append("      catalog:      SHARED_ASSETS.catalog,")
        lines.append("      productSheet: SHARED_ASSETS.productSheet,")
        lines.append("      warranty:     SHARED_ASSETS.warranty,")
        lines.append(f"      tirePhoto:    {img},")
        lines.append("    },")
        lines.append("  },")

    lines.append("];")
    lines.append("")
    lines.append("/** Navbar dropdown section order. */")
    lines.append("export const NAV_GROUP_ORDER: string[] = [")
    for g in NAV_ORDER:
        lines.append(f"  {ts(g)},")
    lines.append("];")
    lines.append("")

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8", newline="\n") as fh:
        fh.write("\n".join(lines))

    print(f"Wrote {os.path.relpath(OUT, ROOT)}  —  {len(out)} tires, "
          f"{sum(len(t['specRows']) for t in out)} spec rows, "
          f"{sum(len(t['features']) for t in out)} feature blocks")

    # Independent count straight off the sheet: every row with a size in it.
    # If this disagrees with what we emitted, the parser dropped something.
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]
    raw, started = 0, False
    for r in range(1, ws.max_row + 1):
        if cell(ws, r, 1).lower() == "tire name" and cell(ws, r, 2):
            started = True
        size = cell(ws, r, 5)
        if started and size and size.lower() != "size":
            raw += 1
    emitted = sum(len(t["specRows"]) for t in out)
    if raw != emitted:
        warnings.append(f"PARSER DROPPED ROWS — the sheet has {raw} size rows but "
                        f"{emitted} were emitted. Do not ship this build.")

    if notes:
        print(f"\n{len(notes)} margin note(s) in column A (kept, not published):")
        for n in notes:
            print("  · " + n)

    if typo_log:
        total = sum(n for _, _, n in typo_log)
        print(f"\nRepaired {total} missing-space-before-'and' typos from the workbook:")
        for tire, field, n in typo_log:
            print(f"  {tire:<20} {field}  x{n}")

    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print("  ! " + w)


if __name__ == "__main__":
    main()
